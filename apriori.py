import pandas as pd
from mlxtend.preprocessing import TransactionEncoder
from mlxtend.frequent_patterns import apriori, association_rules
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.filters import AutoFilter

# Remove strings que não deveriam aparecer
def limpar_arquivo(path, nome_planilha, colunas_para_limpar):
    df = pd.read_excel(path, sheet_name=nome_planilha)
    
    # limpa as colunas específicas
    for coluna in colunas_para_limpar:
        if coluna in df.columns:
            df[coluna] = df[coluna].astype(str).str.replace(
                r"frozenset\(\{'?|'}?\)", "", regex=True
            )
    
    return df

path = 'C:/Users/SDE/Desktop/Codes/Python/Apriori/apriori.xlsx'
df = pd.read_excel(path)

# Agrupar os segmentos por Nro. Único
df_agrupado = df.groupby(df.iloc[:,0])[df.columns[1]].apply(list).reset_index()

# TransactionEncoder para transformar a lista em uma matriz binária (pré-requisito do algoritmo Apriori)
te = TransactionEncoder()
df_transformado = te.fit_transform(df_agrupado.iloc[:,1])
df_binario = pd.DataFrame(df_transformado, columns=te.columns_)
print(df_binario)

# Aplicação do algoritmo apriori
itemset_frequente = apriori(df_binario, min_support=0.01, use_colnames=True)

# Geração das Regras de associação
regra = association_rules(itemset_frequente, metric="confidence", min_threshold=0.7)

# Print dos resultados
print(itemset_frequente)
print(regra)

# Troca os hearders das duas planilhas
itemset_frequente.columns = ["Suporte","Itemset"]


regra.columns = [
    "Conjunto de Produtos da Nota",  # Old 'antecedents'
    "Item a ser ofertado",           # Old 'consequents'
    "Qtd. Notas com o Conjunto de Produtos (Suporte do Antecedente)",  # Old 'antecedent support'
    "Qtd. Notas com o item ofertado (Suporte do Consequente)",         # Old 'consequent support'
    "Qtd. Notas com o conjunto + item ofertado (Suporte da Regra)",    # Old 'support'
    "Perc. Certeza da venda do item (Confiança)",                     # Old 'confidence'
    "Quantas vezes mais chance de vender o item ofertado com este conjunto (Elevação)",  # Old 'lift'
    "Quantas vezes mais o item tem chance de aparecer com o conjunto na nota (Convicção)",  # Old 'conviction'
    "Mede se o conjunto impacta a presença do item (quanto mais próximo de 1 maior) (Grau de Corr.)",  # Old 'leverage'
    "Outra métrica para correlação do conjunto com o item (Métrica de Zhang)"  # Old 'zhangs_metric'
]

with pd.ExcelWriter("resultado_apriori.xlsx", engine='openpyxl') as writer:
    # Salvando os itemsets
    itemset_frequente.to_excel(writer, sheet_name="Itemset", index=False)
    
    # Salvando as regras de associação com as métricas de avaliação de resultado
    regra.to_excel(writer, sheet_name="Regras de Associação", index=False)


colunas_para_limpar = ["Conjunto de Produtos da Nota", "Item a ser ofertado"]



# Limpa os dados com a função para ser usada do workbook
df_limpo = limpar_arquivo("resultado_apriori.xlsx", "Regras de Associação", colunas_para_limpar)

# Carrega o workbook do openpyxl
workbook = load_workbook("resultado_apriori.xlsx")
sheet = workbook["Regras de Associação"]

#Limpa os dados na planilha (com exceção do cabeçalho)
for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
    for cell in row:
        cell.value = None

# Reescreve os dados tratados na planilha
for r_idx, row in enumerate(df_limpo.itertuples(index=False), start=2):  # A partir da segunda linha (para ignorar o cabeçalho)
    for c_idx, value in enumerate(row, start=1):  # A partir da primeira coluna
        sheet.cell(row=r_idx, column=c_idx, value=value)

# Reaplica a formatação
for nome_planilha in ["Itemset", "Regras de Associação"]:
    planilha = workbook[nome_planilha]
    
    # Adiciona o fundo amarelo no cabeçalho
    preenche_header = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    for celula in planilha[1]:  
        celula.fill = preenche_header

    # Adiciona os filtros na coluna
    planilha.auto_filter.ref = planilha.dimensions

    # Formatação para os dados das colunas especificas serem apresentados como porcentagem.
    coluna_porcentagem = [
        "Qtd. Notas com o Conjunto de Produtos (Suporte do Antecedente)",
        "Qtd. Notas com o item ofertado (Suporte do Consequente)",
        "Qtd. Notas com o conjunto + item ofertado (Suporte da Regra)"
    ]
    linha_header = list(planilha[1])
    indice_coluna = {celula.value: celula.column for celula in linha_header}
    for col_nome in coluna_porcentagem:
        if col_nome in indice_coluna:
            col_idx = indice_coluna[col_nome]
            for linha in planilha.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                for celula in linha:
                    celula.number_format = "0.00%"

# Salva o workbook com todas as alterações
workbook.save("resultado_apriori.xlsx")
print("Resultado exportado para 'resultado_apriori.xlsx'")